# 本程序主要用于提取合同日期筛选和发货日期筛选

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import time
import os
import xlrd

# today = datetime(2024,2,23).date()
today = datetime.now().date()  # 获取今日的日期
start_time = time.time()  # 获取程序开始时间


def create_folder(folder_path):
    # exists 函数主要用于判断对象是否存在，而不是获取对象本身。
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)


folder_path = "订单"
create_folder(folder_path)

PATH_1 = "./data.xlsx"  # 设置读取数据位置
PATH_2 = "./data.xlsx"  # 设置读取数据位置
SAVE_PATH = f"./{folder_path}"  # 设置存储数据位置
EXCEL_FILE = f"/{today}_确认订单.xlsx"  # 设置读取数据文件名
FILE_PATH = SAVE_PATH + EXCEL_FILE  # 设置文件路径
COLUMN_1 = {"金开瑞订单编号": "订单编号"}  # 替换行标签名
COLUMN_2 = {"华美订单编号": "订单编号"}  # 替换行标签名
SELECT = "合同日期"  # 选择需要筛选的列
# SELECT = "发货日期"
today_str = SELECT + today.strftime("%Y-%m-%d")  # 设置存储的工作表名称,将日期格式转化为字符串格式


def setdata(PATH, TI_HUAN):
    df = pd.read_excel(PATH)
    df.rename(columns=TI_HUAN, inplace=True)
    df[SCREEN_STR] = pd.to_datetime(df[SCREEN_STR], errors="coerce")
    df = df.dropna(subset=[SCREEN_STR])
    result_df = df[(df[SCREEN_STR].dt.date == today) & (df["分组"] != "质控")]
    return result_df


df1 = pd.read_excel(PATH_1, sheet_name="金开瑞订单编号")
df1.rename(columns=COLUMN_1, inplace=True)  # 替换行标签名
df1[SELECT] = pd.to_datetime(df1[SELECT], errors="coerce")  # 筛选，将列转化为时间格式，没转化成功的会成为nat
df1 = df1.dropna(subset=[SELECT])  # 剔除，将含有nat的行剔除
# 条件筛选
result_df1 = df1[(df1[SELECT].dt.date == today) & (df1["分组"] != "质控")]

df2 = pd.read_excel(PATH_2, sheet_name="华美订单编号")
df2.rename(columns=COLUMN_2, inplace=True)
df2[SELECT] = pd.to_datetime(df2[SELECT], errors="coerce")
df2 = df2.dropna(subset=[SELECT])
result_df2 = df2[(df2[SELECT].dt.date == today) & (df2["分组"] != "质控")]

result = pd.concat([result_df1, result_df2])

# result_df.to_excel( f"C:/Users/Administrator/Desktop/其它应用/today_确认订单/{today}_确认订单.xlsx",encoding = "utf-8",index = False)

df = result
print(f"{today_str}共有： {df.shape[0]} 个。")
# with pd.ExcelWriter(f"{SAVE_PATH}\{today}_确认订单.xlsx", engine='openpyxl') as writer:
with pd.ExcelWriter(FILE_PATH, engine='openpyxl') as writer:
    df.to_excel(writer, today_str, index=False)
    writer.sheets[today_str].column_dimensions["d"].width = 11
    #    writer.sheets["test"]["b2"].number_format = "YYYY-MM-DD" 可以实现单个单元格的格式更改
    for cell in writer.sheets[today_str]["d"]:
        cell.number_format = "YYYY-MM-DD"
    for cell in writer.sheets[today_str]["g"]:
        cell.number_format = "YYYY-MM-DD"
    for cell in writer.sheets[today_str]["w"]:
        cell.number_format = "YYYY-MM-DD"

    # wb = load_workbook(f"{SAVE_PATH}\{today}_确认订单.xlsx")
wb = load_workbook(FILE_PATH)
ws = wb.active

# 将所有使用的行高度设置为20
for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 20

# 将设置列宽
ws.column_dimensions["a"].width = 10
ws.column_dimensions["b"].width = 11
ws.column_dimensions["c"].width = 8
ws.column_dimensions["d"].width = 11
ws.column_dimensions["e"].width = 3
ws.column_dimensions["f"].width = 3
ws.column_dimensions["g"].width = 11
ws.column_dimensions["h"].width = 9
ws.column_dimensions["i"].width = 6
ws.column_dimensions["j"].width = 5
ws.column_dimensions["k"].width = 9
ws.column_dimensions["l"].width = 4
ws.column_dimensions["m"].width = 30
ws.column_dimensions["n"].width = 12
ws.column_dimensions["o"].width = 5
ws.column_dimensions["w"].width = 11

wb.save(FILE_PATH)

end_time = time.time()
ececution_time = round(end_time - start_time, 1)

print(f"用时：{ececution_time} 秒")


